import os
import json
import datetime
import openpyxl


filePath = "./yes24_json_dump_edu/"
wb = openpyxl.Workbook()
now_int_year = datetime.datetime.now().strftime("%Y")
now_int_month = datetime.datetime.now().strftime("%m")
now_int_day = datetime.datetime.now().strftime("%d")
excelfilename = (
    "yes24_edu_" + now_int_year + "_" + now_int_month + "_" + now_int_day + ".xlsx"
)

for _file in os.listdir(filePath):
    with open(filePath + _file, encoding="utf-8") as json_file:
        json_data = json.load(json_file)
        if json_data is None:
            print("파일이 없습니다.")
            exit(1)

    ws = wb.create_sheet(_file[6:10] + _file[11:15])
    ws.cell(column=1, row=1, value="순위")
    ws.cell(column=2, row=1, value="판매지수")
    ws.cell(column=3, row=1, value="제목")
    ws.cell(column=4, row=1, value="가격")
    ws.cell(column=5, row=1, value="저자")
    ws.cell(column=6, row=1, value="출판사")
    ws.cell(column=7, row=1, value="출간일")
    ws.cell(column=8, row=1, value="쪽")
    ws.cell(column=9, row=1, value="카테고리+태그")
    ws.cell(column=10, row=1, value="ISBN")
    ws.cell(column=11, row=1, value="URL")

    row = 2

    for key, value in json_data.items():
        title = value["title"]
        price = value["right_price"]
        author = ",".join(value["author"])
        publisher = value["publisher"]
        date_of_publish = value["publish_date"]
        rank = value["rank"]
        selling_point = value["sales_point"]
        page = value["page"]
        category_of_yes24 = ",".join(value["tags"])
        isbn = value["isbn"]
        url = value["url"]

        ws.cell(column=1, row=row, value=rank)
        ws.cell(column=2, row=row, value=selling_point)
        ws.cell(column=3, row=row, value=title)
        ws.cell(column=4, row=row, value=price)
        ws.cell(column=5, row=row, value=author)
        ws.cell(column=6, row=row, value=publisher)
        ws.cell(column=7, row=row, value=date_of_publish)
        ws.cell(column=8, row=row, value=page)
        ws.cell(column=9, row=row, value=category_of_yes24)
        ws.cell(column=10, row=row, value=isbn)
        ws.cell(column=11, row=row, value=url)

        row = row + 1

wb.save(filename=excelfilename)