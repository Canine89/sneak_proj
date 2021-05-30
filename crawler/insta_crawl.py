from selenium import webdriver
from time import sleep
from openpyxl import Workbook
import json
import datetime


# Json 파일 껍데기 생성(템플릿)
def toJson(insta_dict, file_name):
    with open(
        file_name + str(datetime.datetime.now()) + ".json", "w", encoding="utf-8"
    ) as file:
        # 정의해 놓은 Json 파일에 크롤링한 데이터 넣어
        json.dump(insta_dict, file, ensure_ascii=False, indent="\t")
    # try:
    #     with open(file_name + '.json', 'w', encoding='utf-8') as file:
    #         # 정의해 놓은 Json 파일에 크롤링한 데이터 넣어
    #         json.dump(insta_dict, file, ensure_ascii=False, indent='\t')
    # except:
    #     print("something wrong...")
    #     return -1

    return 0


if __name__ == "__main__":
    # chrome driver 위치
    driver = webdriver.Chrome(executable_path="../webdriver/chromedriver 3")

    # excel 타이틀, 위치 지정
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "인스타그램 해시태그"
    ws1.append(["키워드", "해시태그 수"])
    dest_filename = "./SNS 크롤링.xlsx"

    # 크롤링 1 - 이동, 버튼 누르기
    url = "https://www.instagram.com/"
    driver.get(url)  # 엔터
    sleep(3)  # 호출했는데 데이터가 오기까지 시간이 걸리는 경우가 있으므로 대기 시간 지정

    item = driver.find_element_by_xpath(
        '//*[@id="loginForm"]/div/div[1]/div/label/input'
    )
    item.send_keys("he_e_young")  # id 날리기

    item = driver.find_element_by_xpath(
        '//*[@id="loginForm"]/div/div[2]/div/label/input'
    )
    item.send_keys("cjstk2137^")  # pass 날리기

    item = driver.find_element_by_xpath('//*[@id="loginForm"]/div/div[3]/button/div')
    item.click()  # 버튼 클릭

    sleep(5)

    # 크롤링 2 키워드 입력, 해시태그 긁어 오기
    keywords = [
        "프리미어프로",
        "파이썬",
        "이지스퍼블리싱",
    ]
    _row = 1
    hashList = list()
    for keyword in keywords:
        temp = dict()  # 딕셔너리로 저장

        url = "https://www.instagram.com/explore/tags/" + keyword + "/"
        driver.get(url)  # 엔터
        sleep(5)
        hash_tag_count = driver.find_element_by_xpath(
            "/html/body/div[1]/section/main/header/div[2]/div/div[2]/span/span"
        ).text

        # 엑셀로 데이터 저장
        temp["해시태그 수"] = hash_tag_count  # 해시태그 수 : 숫자 딕셔너리로 조합
        ws1.cell(row=_row, column=2, value=temp["해시태그 수"])

        print(keyword)  # 키워드
        temp["키워드"] = keyword  # 키워드 : 키워드 값
        ws1.cell(row=_row, column=3, value=temp["키워드"])
        _row = _row + 1

        hashList.append(temp)

    # hashList를 가지고 아까 시켜 둔 거 있지? 그거 해
    toJson(insta_dict=hashList, file_name="인별그램")
    # result = toJson(hashList, '인별그램')
    # if result == 0:
    #     print("well crawling...")
    # elif result == -1:
    #     print("bad crawling...")

    # driver 종료
    driver.close()
